from openpyxl import load_workbook
from datetime import datetime
from dateutil import parser
from matplotlib.dates import date2num
import os, csv, random, pandas as pd, matplotlib.pyplot as plt, matplotlib.dates as mdates, bcrypt

#TESTED
def addTransactionDetails(accno, amount, details):

    wb = load_workbook(filename="MNMBankDatabase.xlsx")
    write = pd.ExcelWriter('MNMBankDatabase.xlsx', engine = 'openpyxl', mode='a')
    write.book = wb
    write.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name=accno)
    dat.to_csv('addTransaction.csv', index=None, header=True)

    balance = fetchBalance(accno)
    now = datetime.now()
    # dd/mm/YY H:M:S
    dt_string = now.strftime("%d-%m-%Y %H:%M:%S")

    with open('addTransaction.csv', 'a', newline="") as f:
        writer = csv.writer(f)
        if amount < 0:
            writer.writerow([dt_string, details, 0, -amount, int(balance)+amount])
        else:
            writer.writerow([dt_string, details, amount, 0, int(balance)+amount])

    read_file = pd.read_csv('addTransaction.csv')
    read_file.to_excel(write, sheet_name = accno, index = None, header=True)
    write.save()
    write.close()
    os.remove('addTransaction.csv')

#TESTED
def addAccount(name, dob, aadharno, panno, address, email, contact, password, balance):

    wb = load_workbook(filename="MNMBankDatabase.xlsx")
    write = pd.ExcelWriter('MNMBankDatabase.xlsx', engine = 'openpyxl', mode='a')
    write.book = wb
    write.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    #Current date&time
    now = datetime.now()
    # dd/mm/YY H:M:S
    dt_string = now.strftime("%d-%m-%Y %H:%M:%S")
    accno = generateAccountNo()

    #Writing personal details to its sheet
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name='Personal Details')
    dat.to_csv('addAccountPersonal.csv', index=None, header=True)

    with open('addAccountPersonal.csv', 'a', newline="") as f:
        writer = csv.writer(f)
        writer.writerow([accno, name, dob, aadharno, panno, email, contact, address])
    read_file = pd.read_csv('addAccountPersonal.csv')
    read_file.to_excel(write, sheet_name = 'Personal Details', index = None, header=True)

    #Creating an account sheet
    with open('addAccountSheet.csv', 'w', newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Date_Time", "Note", "Credit", "Debit", "Balance"])
        writer.writerow([dt_string, "Initial Balance", balance, 0, balance])
    read_file = pd.read_csv('addAccountSheet.csv')
    read_file.to_excel(write, sheet_name = accno, index = None, header=True)

    #Entering accno&password to its sheet
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name='Login Details')
    dat.to_csv('addAccountSheet.csv', index=None, header=True)

    with open('addAccountSheet.csv', 'a', newline="") as f:
        writer = csv.writer(f)
        writer.writerow([accno, generatePasswordHash(password)])
    read_file = pd.read_csv('addAccountSheet.csv')
    read_file.to_excel(write, sheet_name = 'Login Details', index = None, header=True)

    write.save()
    write.close()
    os.remove('addAccountPersonal.csv')
    os.remove('addAccountSheet.csv')
    return accno

#TESTED
def generatePasswordHash(password):
    b_pwd = bytes(password, encoding = 'utf-8')
    hash = bcrypt.hashpw(b_pwd, bcrypt.gensalt())
    return hash

#TESTED
def checkPasswordHash(password, hash):
    b_pwd = bytes(password, encoding = 'utf-8')
    b_hash = bytes(hash[2:-1], encoding = 'utf-8')
    if bcrypt.checkpw(b_pwd, b_hash):
        return True
    return False

#TESTED
def checkLogin(accno, password):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name='Login Details')
    dat.to_csv('checkLoginDetails.csv', index=None, header=True)

    flag = 0
    with open('checkLoginDetails.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if row[0] == accno and checkPasswordHash(password, row[1]):
                flag = 1
                break
    os.remove('checkLoginDetails.csv')
    if flag == 1:
        return True
    else:
        return False

#TESTED
def fetchBalance(accno):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name=accno)
    dat.to_csv('fetchBal.csv', index=None, header=True)

    with open('fetchBal.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            balance = row[-1]
    os.remove('fetchBal.csv')
    return balance

#TESTED
def viewPersonalDetails(accno):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name='Personal Details')
    dat.to_csv('viewPersonalDetails.csv', index=None, header=True)

    details = []
    with open('viewPersonalDetails.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if row[0] == accno:
                details = row
                break
    os.remove('viewPersonalDetails.csv')
    return details

#TESTED
def viewTransactionDetails(accno):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name=accno)
    dat.to_csv('viewTransaction.csv', index=None, header=True)

    details = []
    with open('viewTransaction.csv', 'r') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            details.append(row)
    os.remove('viewTransaction.csv')
    return details

#TESTED
def generateAccountNo():

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name='Personal Details')
    dat.to_csv('generateAcNum.csv', index=None, header=True)

    acnum = random.randint(62501000000, 62502000000)
    flag = 0

    with open('generateAcNum.csv', 'r') as f:
        reader = csv.reader(f)
        while True:
            for row in reader:
                if row[0] == acnum:
                    flag = 1
                    break
            if flag == 0:
                break
            acnum = random.randint(62501000000, 62502000000)

    os.remove('generateAcNum.csv')
    return str(acnum)

#TESTED
def generateLoanNo():

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name='Loan Details')
    dat.to_csv('generateLNum.csv', index=None, header=True)

    loannum = random.randint(4180100000, 4180200000)
    flag = 0

    with open('generateLNum.csv', 'r') as f:
        reader = csv.reader(f)
        while True:
            for row in reader:
                if row[1] == loannum:
                    flag = 1
                    break
            if flag == 0:
                break
            loannum = random.randint(4180100000, 4180200000)

    os.remove('generateLNum.csv')
    return str(loannum)

#TESTED
def generateFDNo():

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name='FD Details')
    dat.to_csv('generateFdNum.csv', index=None, header=True)

    fdnum = random.randint(2850100000, 2850200000)
    flag = 0

    with open('generateFdNum.csv', 'r') as f:
        reader = csv.reader(f)
        while True:
            for row in reader:
                if row[1] == fdnum:
                    flag = 1
                    break
            if flag == 0:
                break
            fdnum = random.randint(2850100000, 2850200000)

    os.remove('generateFdNum.csv')
    return str(fdnum)

#TESTED
def takeLoan(accno, type, time_period, amount):

    wb = load_workbook(filename="MNMBankDatabase.xlsx")
    write = pd.ExcelWriter('MNMBankDatabase.xlsx', engine = 'openpyxl', mode='a')
    write.book = wb
    write.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    #Creating a Loan Account
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name="Loan Details")
    dat.to_csv('createLoanDetails.csv', index=None, header=True)

    loan_num = generateLoanNo()
    with open('createLoanDetails.csv', 'a', newline="") as f:
        writer = csv.writer(f)
        writer.writerow([accno, loan_num, type, 0, int(amount*(1+(0.08*time_period))), time_period, int(amount*(1+(0.08*time_period)))])
    read_file = pd.read_csv('createLoanDetails.csv')
    read_file.to_excel(write, sheet_name = "Loan Details", index = None, header=True)

    #Updating Account Passbook
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name=accno)
    dat.to_csv('UpdatePassbookNewLoan.csv', index=None, header=True)

    balance = fetchBalance(accno)
    now = datetime.now()
    # dd/mm/YY H:M:S
    dt_string = now.strftime("%d-%m-%Y %H:%M:%S")

    with open('UpdatePassbookNewLoan.csv', 'a', newline="") as f:
        writer = csv.writer(f)
        writer.writerow([dt_string, f"New Loan: {loan_num}", amount, 0, int(balance)+amount])

    read_file = pd.read_csv('UpdatePassbookNewLoan.csv')
    read_file.to_excel(write, sheet_name = accno, index = None, header=True)
    write.save()
    write.close()
    os.remove('UpdatePassbookNewLoan.csv')
    os.remove('createLoanDetails.csv')

    return loan_num

#TESTED
def checkLoan(accno, loanno):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name="Loan Details")
    dat.to_csv('loanCheck.csv', index=None, header=True)

    flag = 0
    with open('loanCheckDetails.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if row[0] == accno and row[1] == loanno:
                flag = 1
                break
    os.remove('loanCheckDetails.csv')
    if flag == 1:
        return True
    return False

#TESTED
def checkSufficientBalance(accno, amount):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name=accno)
    dat.to_csv('checkSufficientBalance.csv', index=None, header=True)

    with open('checkSufficientBalance.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            balance = row[-1]

    os.remove('checkSufficientBalance.csv')
    if int(balance) >= amount:
        return True
    return False

#TESTED
def payLoan(accno, loanno, amount):

    wb = load_workbook(filename="MNMBankDatabase.xlsx")
    write = pd.ExcelWriter('MNMBankDatabase.xlsx', engine = 'openpyxl', mode='a')
    write.book = wb
    write.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    #Modifying Loan Account
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name="Loan Details")
    dat.to_csv('payLoanDetails.csv', index=None, header=True)

    loan_amt_remaining = 0
    with open('payLoanDetails.csv', 'r') as f, open('new_payLoanDetails.csv', 'w') as f2:
        reader = csv.reader(f)
        writer = csv.writer(f2)
        for row in reader:
            if row[0] == accno and row[1] == loanno:
                loan_amt_remaining = int(float(row[4]))-amount
                writer.writerow([accno, loanno, row[2], int(float(row[3]))+amount, loan_amt_remaining, row[-2], row[-1]])
            else:
                writer.writerow(row)

    read_file = pd.read_csv('new_payLoanDetails.csv')
    read_file.to_excel(write, sheet_name = "Loan Details", index = None, header=True)
    os.remove('payLoanDetails.csv')
    os.remove('new_payLoanDetails.csv')

    #Updating Account Passbook
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name=accno)
    dat.to_csv('UpdatePassbookPayLoan.csv', index=None, header=True)

    balance = fetchBalance(accno)
    now = datetime.now()
    # dd/mm/YY H:M:S
    dt_string = now.strftime("%d-%m-%Y %H:%M:%S")

    with open('UpdatePassbookPayLoan.csv', 'a', newline="") as f:
        writer = csv.writer(f)
        writer.writerow([dt_string, f"Loan Installment: {loanno}", 0, amount, int(balance)-amount])

    read_file = pd.read_csv('UpdatePassbookPayLoan.csv')
    read_file.to_excel(write, sheet_name = accno, index = None, header=True)
    os.remove('UpdatePassbookPayLoan.csv')

    write.save()
    write.close()

    return loan_amt_remaining

#TESTED
def viewLoanDetails(accno):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name="Loan Details")
    dat.to_csv('viewloandets.csv', index=None, header=True)

    details = []
    with open('viewloandets.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if accno == row[0] and int(float(row[-3])) > 0:
                details.append(row[1:])
    os.remove('viewloandets.csv')
    return details

#TESTED
def createFD(accno, maturity_period, amount):

    wb = load_workbook(filename="MNMBankDatabase.xlsx")
    write = pd.ExcelWriter('MNMBankDatabase.xlsx', engine = 'openpyxl', mode='a')
    write.book = wb
    write.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    #Creating a FD Account
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name="FD Details")
    dat.to_csv('createFDDetails.csv', index=None, header=True)

    fd_num = generateFDNo()
    with open('createFDDetails.csv', 'a', newline="") as f:
        writer = csv.writer(f)
        writer.writerow([accno, fd_num, maturity_period, 6, amount, int(amount*(1+(0.06*maturity_period))), 'No'])
    read_file = pd.read_csv('createFDDetails.csv')
    read_file.to_excel(write, sheet_name = "FD Details", index = None, header=True)

    #Updating Account Passbook
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name=accno)
    dat.to_csv('UpdatePassbookNewFD.csv', index=None, header=True)

    balance = fetchBalance(accno)
    now = datetime.now()
    # dd/mm/YY H:M:S
    dt_string = now.strftime("%d-%m-%Y %H:%M:%S")

    with open('UpdatePassbookNewFD.csv', 'a', newline="") as f:
        writer = csv.writer(f)
        writer.writerow([dt_string, f"New FD: {fd_num}", 0, amount, int(balance)-amount])

    read_file = pd.read_csv('UpdatePassbookNewFD.csv')
    read_file.to_excel(write, sheet_name = accno, index = None, header=True)
    write.save()
    write.close()
    os.remove('UpdatePassbookNewFD.csv')
    os.remove('createFDDetails.csv')

    return fd_num

#TESTED
def viewFDDetails(accno):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name="FD Details")
    dat.to_csv('viewfddets.csv', index=None, header=True)

    details = []
    with open('viewfddets.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if accno == row[0] and row[-1] == "No":
                details.append(row[1:-1])
    os.remove('viewfddets.csv')
    return details

#TESTED
def checkFD(accno, fdno):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name="FD Details")
    dat.to_csv('FDCheckDetails.csv', index=None, header=True)

    flag = 0
    with open('FDCheckDetails.csv', 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            if row[0] == accno and row[1] == fdno:
                flag = 1
                break
    os.remove('FDCheckDetails.csv')
    if flag == 1:
        return True
    return False

#TESTED
def breakFD(accno, fdno):

    wb = load_workbook(filename="MNMBankDatabase.xlsx")
    write = pd.ExcelWriter('MNMBankDatabase.xlsx', engine = 'openpyxl', mode='a')
    write.book = wb
    write.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    #Modifying Loan Account
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name="FD Details")
    dat.to_csv('breakFD.csv', index=None, header=True)

    amount = 0
    with open('breakFD.csv', 'r') as f, open('new_breakFD.csv', 'w') as f2:
        reader = csv.reader(f)
        writer = csv.writer(f2)
        for row in reader:
            if row[0] == accno and row[1] == fdno:
                amount = int(float(row[-3]))
                writer.writerow([accno, fdno, row[2], row[3], row[4], row[5], "Yes"])
            else:
                writer.writerow(row)

    read_file = pd.read_csv('new_breakFD.csv')
    read_file.to_excel(write, sheet_name = "FD Details", index = None, header=True)
    os.remove('breakFD.csv')
    os.remove('new_breakFD.csv')

    #Updating Account Passbook
    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name=accno)
    dat.to_csv('UpdatePassbookBreakFD.csv', index=None, header=True)

    balance = fetchBalance(accno)
    now = datetime.now()
    # dd/mm/YY H:M:S
    dt_string = now.strftime("%d-%m-%Y %H:%M:%S")

    with open('UpdatePassbookBreakFD.csv', 'a', newline="") as f:
        writer = csv.writer(f)
        writer.writerow([dt_string, f"Break FD: {fdno}", amount, 0, int(balance)+int(amount)])

    read_file = pd.read_csv('UpdatePassbookBreakFD.csv')
    read_file.to_excel(write, sheet_name = accno, index = None, header=True)
    os.remove('UpdatePassbookBreakFD.csv')

    write.save()
    write.close()

#TESTED
def viewTransactionGraph(accno):

    dat = pd.read_excel('MNMBankDatabase.xlsx', sheet_name=accno)
    dat.to_csv('graph.csv', index=None, header=True)

    dates, balance = [], []
    with open('graph.csv', 'r') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            date = datetime.strptime(row[0][:10], "%d-%m-%Y")
            dates.append(date)
            balance.append(row[-1])

    date_bal_dict = {}
    for i in range(len(dates)):
        date_bal_dict[dates[i]] = balance[i]
    new_dates = list(date_bal_dict.keys())
    new_balance = list(date_bal_dict.values())
    os.remove('graph.csv')

    plt.gcf().autofmt_xdate()
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    plt.gca().xaxis.set_major_locator(mdates.DayLocator())
    plt.plot_date(new_dates, new_balance, linestyle='solid')
    plt.show()

#print(checkSufficientBalance('6250', 20000))
#payLoan('6250', '4180171721', 200)
#print(fetchBalance('6250'))
#print(viewLoanDetails('6250', '4180171721'))
#print(viewFDDetails('6250', '2850130065'))
##createFD("6250", 5, 20000)
#print(fetchBalance('6250'))
#viewTransactionGraph('6250')
